# # import pandas as pd
# # import os

# # # Read SSN list correctly
# # series = pd.read_excel(r"C:\Users\krish\OneDrive\Desktop\SSN.xlsx")
# # ssn_list = series.iloc[:, 1].dropna().astype(str).tolist()  # Assuming SSNs are in the first column

# # # Read the CSV file
# # df = pd.read_csv(r"C:\Users\krish\OneDrive\Desktop\OOE20250228085048.csv", dtype={'SSN': str})  # Ensure SSN is treated as string

# # # Filter DataFrame
# # filtered_df = df[df['SSN'].isin(ssn_list)]

# # # Define output path
# # directory = 'output'
# # filename = 'filtered.csv'
# # path = os.path.join(directory, filename)

# # # Create output directory if it doesn't exist
# # os.makedirs(directory, exist_ok=True)

# # # Save the filtered DataFrame
# # filtered_df.to_csv(path, index=False)

# # print(f"Filtered data saved to {path}")


# # import ast

# # # Replace 'your_file.txt' with the actual file name
# # file_path = r"C:\Users\krish\OneDrive\Desktop\output_db_dict.txt"

# # with open(file_path, "r") as file:
# #     data = file.read()
# # print(len(data))
# # # Convert text to list using ast.literal_eval (safe parsing)
# # if True:
# #     data_list = ast.literal_eval(data)
# #     if isinstance(data_list, list):
# #         print("Length of the list:", len(data_list))
# #     else:
# #         print("The file does not contain a list.")

# from datetime import datetime, timedelta
# import calendar

# # Get today's date
# # today = datetime.today()
# # last_day_prev_month = datetime(today.year, today.month, 1) - timedelta(days=1)
# # formatted_date_last_month = last_day_prev_month.strftime("%m/%d/%Y")

# import pandas as pd

# # # Read the CSV file
# # df = pd.read_csv(r"C:\Users\krish\OneDrive\Desktop\db2_output_data.txt", delimiter=",")  # Change delimiter if needed

# # # Save as Excel
# # df.to_excel("formatted_file.xlsx", index=False)

# # print("Excel file created successfully!")

# # import pandas as pd

# # # Load your dataset
# # df = pd.read_csv(r"C:\Users\krish\OneDrive\Desktop\db2_output_data.txt", delimiter=",")

# # # Check Column 21 (TX1DES) for non-ASCII characters
# # def find_non_ascii(value):
# #     return any(ord(char) > 127 for char in str(value))

# # # Filter rows where TX1DES has non-ASCII characters
# # problematic_rows = df[df["TX1DES"].apply(find_non_ascii)]

# # print(problematic_rows['TX1DES'])



# # <Layout style={{ minHeight: '100vh', backgroundColor: '#FBF6E9' }}>
# #       <Sider width={256} style={{ backgroundColor: '#5DB996' }}>
# #         <div className="logo" style={{ padding: '16px', textAlign: 'center', color: '#FBF6E9' }}>
# #           <h3>Pharmagretech</h3>
# #         </div>
# #         <Menu
# #           theme="dark"
# #           mode="inline"
# #           style={{ backgroundColor: '#5DB996', color: '#FBF6E9' }}
# #           onClick={({ key }) => setCurrentView(key)}
# #           openKeys={openKeys}
# #           onOpenChange={onOpenChange}
# #         >
# #           <Menu.Item key="dashboard" icon={<BarChartOutlined />}>
# #             Dashboard
# #           </Menu.Item>
# #           <Menu.Item key="attendance" icon={<UserOutlined />}>
# #             Attendance
# #           </Menu.Item>
# #           <SubMenu key="payroll" icon={<MoneyCollectOutlined />} title="Payroll" className="Submenuclass">
# #             <Menu.Item key="payments" style={{ backgroundColor: '#5DB996', color: '#FBF6E9' }} icon={<CreditCardOutlined />}>
# #               Payments
# #             </Menu.Item>
# #             <Menu.Item key="bank" style={{ backgroundColor: '#5DB996', color: '#FBF6E9' }} icon={<BankOutlined />}>
# #               Bank Details
# #             </Menu.Item>
# #           </SubMenu>
# #           <Menu.Item key="leave" icon={<CalendarOutlined />}>
# #             Leave
# #           </Menu.Item>
# #           <SubMenu key="employees" icon={<TeamOutlined />} title="Employees" className="Submenuclass">
# #             <Menu.Item key="addEmployee" style={{ backgroundColor: '#5DB996', color: '#FBF6E9' }} icon={<UserAddOutlined />}>
# #               Add Employee
# #             </Menu.Item>
# #             <Menu.Item key="removeEmp" style={{ backgroundColor: '#5DB996', color: '#FBF6E9' }} icon={<UserDeleteOutlined />}>
# #               Remove Employee
# #             </Menu.Item>
# #           </SubMenu>
# #           <SubMenu key="settings" icon={<SettingOutlined />} title="Settings" className="Submenuclass">
# #           <Menu.Item key="profileSet" style={{ backgroundColor: '#5DB996', color: '#FBF6E9' }} icon={<UserOutlined />}>
# #             Profile Settings
# #           </Menu.Item>
# #           <Menu.Item key="employeeSet" style={{ backgroundColor: '#5DB996', color: '#FBF6E9' }} icon={<TeamOutlined />}>
# #             Employee Settings
# #           </Menu.Item>
# #         </SubMenu>
# #           <Menu.Item key="logout" icon={<LogoutOutlined />} onClick={() => router.push('/')}>
# #             Logout
# #           </Menu.Item>
# #         </Menu>
# #       </Sider>

# @api_view(['POST'])
# def add_dependents_db2(request):
#     relationship = request.data.get("relationship")  
#     if relationship == "Member":
#         return Response({"error": "Only Dependents can be added. Invalid relationship type."}, status=400)

#     EMSSN = request.data.get("emssn")  
#     DPDSSN = request.data.get("dpdssn")
#     if not EMSSN:
#         return Response({"error": "SSN (emssn) is required."}, status=400)

#     try:
#         connection = pyodbc.connect(connection_string)
#         cursor = connection.cursor()
        
#         cursor.execute(f"SELECT COUNT(*) FROM {schema_name}.depnp WHERE dpdssn = ?", (EMSSN,))
#         if cursor.fetchone()[0] > 0:
#             return Response({"error": "SSN already exists. Duplicate entries are not allowed."}, status=400)

#         dpname = request.data.get("dpname")
#         dpsex = request.data.get("dpsex")
#         dpdob = request.data.get("dpdob")  
#         dpadr1 = request.data.get("dpadr1")
#         dpcity = request.data.get("dpcity")
#         dpst = request.data.get("dpst")
#         country = request.data.get("country")  
#         dpmem = request.data.get("dpmem")
#         dpplan = request.data.get("dpplan")
#         dpclas = request.data.get("dpclas")

#         dob_parsed = None
#         dpdoby = dpdobm = dpdobd = None
#         if dpdob:
#             try:
#                 dob_parsed = datetime.strptime(dpdob, "%Y-%m-%d")
#                 dpdoby = dob_parsed.year
#                 dpdobm = dob_parsed.month
#                 dpdobd = dob_parsed.day
#             except ValueError:
#                 return Response({"error": "Invalid date format. Use YYYY-MM-DD."}, status=400)

#         full_address = ", ".join(filter(None, [dpadr1, dpcity, dpst, country]))
#         if relationship == "Spouse" or relationship == 'spouse':
#             dprltn = 1
#         elif relationship == "Son" or relationship == 'son':
#             dprltn = 2
#         elif relationship == 'Daughter' or relationship == 'daughter':
#             dprltn = 3
#         elif relationship == 'Stepchild' or relationship == 'stepchild':
#             dprltn = 4
#         else:
#             dprltn = 9

#         current_date = datetime.today()

#         year = str(current_date.year)
#         month = str(current_date.month)
#         day = str(current_date.day)
#         status = 'A'
        
#         insert_query = f"""
#             INSERT INTO {schema_name}.depnp (dpssn, dpname, dpsex, dpdoby, dpdobm, dpdobd, dpdssn, dpclas, dpplan,dprltn,dpefdy,dpefdm,dpefdd,dpstat,dpupyy,dpupmm,dpupdd)
#             VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?) WITH NC """
        
#         cursor.execute(insert_query, (EMSSN, dpname, dpsex, dpdoby, dpdobm, dpdobd, DPDSSN, dpclas, dpplan,dprltn,year,month,day,status,year,month,day))
#         connection.commit()

#         return Response({"message": "Dependent added successfully!", "status": "success"})

#     except pyodbc.Error as e:
#         return Response({"error": str(e)}, status=400)
#     finally:
#         cursor.close()
#         connection.close()

    
# # from rest_framework.views import APIView
# # from rest_framework.response import Response
# # from rest_framework import status
# # import pyodbc
# # from datetime import datetime



# # class GetMemberInfoDB2(APIView):
# #     def get(self, request):
# #         name = request.GET.get('name')
# #         relationship = request.GET.get('relationship')
# #         ssn = request.GET.get('ssn')

# #         if not name and not relationship and not ssn:
# #             return Response({"error": "All fields are required"}, status=status.HTTP_400_BAD_REQUEST)

# #         connection = pyodbc.connect(connection_string)
# #         cursor = connection.cursor()

# #         try:
# #             if relationship.lower() == "member":
# #                 member_query = f"SELECT * FROM {schema_name}.empyp WHERE emssn = ?"
# #                 cursor.execute(member_query, (ssn,))
# #                 member_row = cursor.fetchone()

# #                 if not member_row:
# #                     return Response({"error": "Member not found"}, status=status.HTTP_404_NOT_FOUND)

# #                 member_columns = [desc[0].lower() for desc in cursor.description]
# #                 member = dict(zip(member_columns, member_row))
# #                 print(">>>>>>>>>>>>>>>>>><<<<<<<<<<<<<<<<<<<<<<<")
# #                 print(member.get('emdobm'), member.get('emdobd'), member.get('emdoby'))

# #                 dob = self.format_dob(member.get('emdobm'), member.get('emdobd'), member.get('emdoby'))

# #                 data = {
# #                     "name": member.get('emname', "name is not available"),
# #                     "ssn": member.get('emssn', "SSN is not available"),
# #                     "relationship": relationship,
# #                     "member_id": member.get('emmem#', "member ID is not available"),
# #                     "dob": dob if dob else "DOB is not available",
# #                     "address": member.get('emadr1', "address is not available"),
# #                     "state": member.get('emst', "state is not available"),
# #                     "city": member.get('emcity', "city is not available"),
# #                     "country": "USA"
# #                 }

# #             else:
# #                 dependent_query = f"SELECT * FROM {schema_name}.depnp WHERE dpdssn = ? AND dpname LIKE ?"
# #                 cursor.execute(dependent_query, (ssn, f"%{name}%"))
# #                 dependent_row = cursor.fetchone()

# #                 if not dependent_row:
# #                     return Response({"error": "Dependent not found"}, status=status.HTTP_404_NOT_FOUND)

# #                 dependent_columns = [desc[0].lower() for desc in cursor.description]
# #                 dependent = dict(zip(dependent_columns, dependent_row))

# #                 dob = self.format_dob(dependent.get('dpdobm'), dependent.get('dpdobd'), dependent.get('dpdoby'))

# #                 member_query = f"SELECT * FROM {schema_name}.empyp WHERE emssn = ?"
# #                 cursor.execute(member_query, (dependent.get('dpssn'),))
# #                 member_row = cursor.fetchone()
# #                 member = dict(zip([desc[0].lower() for desc in cursor.description], member_row)) if member_row else {}

# #                 data = {
# #                     "name": dependent.get('dpname', "name is not available"),
# #                     "ssn": dependent.get('dpdssn', "SSN is not available"),
# #                     "relationship": dependent.get('dptype', "relationship is not available"),
# #                     "member_id": member.get('emmem', "member ID is not available"),
# #                     "dob": dob if dob else "DOB is not available",
# #                     "address": member.get('emadr1', "address is not available"),
# #                     "state": member.get('emst', "state is not available"),
# #                     "city": member.get('emcity', "city is not available"),
# #                     "country": "USA"
# #                 }

# #         finally:
# #             cursor.close()
# #             connection.close()

# #         return Response(data, status=status.HTTP_200_OK)

# #     def format_dob(self, month, day, year):
# #         if not all([month, day, year]):
# #             return None

# #         try:
# #             # Convert Decimal to int
# #             month = int(month) if isinstance(month, Decimal) else month
# #             day = int(day) if isinstance(day, Decimal) else day
# #             year = int(year) if isinstance(year, Decimal) else year

# #             dob = datetime(year, month, day)
# #             return dob.strftime("%B %d, %Y")
# #         except (ValueError, TypeError):
# #             return None

    
        
# # class GetMemberInfo(APIView):
# #     def get(self, request):
# #         name = request.GET.get('name')
# #         relationship = request.GET.get('relationship')
# #         ssn = request.GET.get('ssn')

# #         if not name and not relationship and not ssn:
# #             return Response("All fields are required")
        
# #         if relationship.lower() == "member":
# #             member = MyappEmpyp.objects.filter(emssn=ssn).first()
# #             if not member:
# #                 return Response({'error': 'Member not found'}, status=404)
            
# #             dob = self.format_dob(member.emdobm, member.emdobd, member.emdoby)
        
# #             data = {
# #                 "name": member.emname if member.emname else "name is not available",
# #                 "ssn": member.emssn if member.emssn else "SSN is not available",
# #                 "relationship": relationship,
# #                 "member_id": member.emmem if member.emmem else "member ID is not available",
# #                 "dob": dob if dob else "DOB is not available",
# #                 "address": member.emadr1 if member.emadr1 else "address is not available",
# #                 "state": member.emst if member.emst else "state is not available",
# #                 "city": member.emcity if member.emcity else "city is not available",
# #                 "country": "USA"
# #             } 
        
# #         else:
# #             dependent = MyappDepnp.objects.filter(dpdssn=ssn, dpname__icontains=name).first()
# #             if not dependent:
# #                 return Response({"error": "Dependent not found"}, status=404)

# #             dob = self.format_dob(dependent.dpdobm, dependent.dpdobd, dependent.dpdoby)

# #             member = MyappEmpyp.objects.filter(emssn=dependent.dpssn).first()

# #             data = {
# #                 "name": dependent.dpname if dependent.dpname else "name is not available",
# #                 "ssn": dependent.dpdssn if dependent.dpdssn else "SSN is not available",
# #                 "relationship": dependent.dptype if dependent.dptype else "relationship is not available",
# #                 "member_id": member.emmem if member and member.emmem else "member ID is not available",
# #                 "dob": dob if dob else "DOB is not available",
# #                 "address": member.emadr1 if member and member.emadr1 else "address is not available",
# #                 "state": member.emst if member and member.emst else "state is not available",
# #                 "city": member.emcity if member and member.emcity else "city is not available",
# #                 "country": "USA"
# #             }

# #         return Response(data)      
    
# #     def format_dob(self, month, day, year):
# #         if not (month and day and year):
# #             return None
# #         try:
# #             dob = datetime(year, month, day)
# #             return dob.strftime("%B %d, %Y")
# #         except ValueError:
# #             return None
        

# # class UpdateMemberInfo(APIView):
# #     def post(self, request):
# #         name = request.data.get('name')
# #         relationship = request.data.get('relationship')
# #         ssn = request.data.get('ssn')
# #         member_id = request.data.get('member_id')
# #         dob_str = request.data.get('dob')  
# #         address = request.data.get('address')
# #         state = request.data.get('state')
# #         city = request.data.get('city')

# #         if not all([name, relationship, ssn, member_id, dob_str, address, state, city]):
# #             return Response({"error": "Missing required fields"}, status=400)

# #         year, month, day = self.parse_dob(dob_str)
# #         if not all([year, month, day]):
# #             return Response({"error": "Invalid date format. Expected format like 'mm-dd-yyyy'"}, status=400)

# #         if relationship.lower() == "member":
# #             instance = MyappEmpyp.objects.filter(emssn=ssn).first()
# #             if not instance:
# #                 return Response({"error": "Member record not found"}, status=404)

# #             instance.emname = name
# #             instance.emssn = ssn
# #             instance.emmem = member_id
# #             instance.emdoby = year
# #             instance.emdobm = month
# #             instance.emdobd = day
# #             instance.emadr1 = address
# #             instance.emst = state
# #             instance.emcity = city
# #             instance.save()

# #             return Response({"message": "Member record updated successfully"})

# #         else:
# #             instance = MyappDepnp.objects.filter(dpdssn=ssn, dpname__icontains=name).first()
# #             if not instance:
# #                 return Response({"error": "Dependent record not found"}, status=404)

# #             with transaction.atomic():
# #                 rows_updated = MyappDepnp.objects.filter(id=instance.id).update(
# #                     dpname=name,
# #                     dpdssn=ssn,
# #                     dptype=relationship,
# #                     dpdoby=year,
# #                     dpdobm=month,
# #                     dpdobd=day,
# #                 )

# #             if rows_updated == 0:
# #                 return Response({"error": "Update failed!"}, status=500)
            
# #             return Response({"message": "Dependent record updated successfully"})

# #     def parse_dob(self, dob_str):
# #         try:
# #             dt = datetime.strptime(dob_str, "%m-%d-%Y") 
# #             return dt.year, dt.month, dt.day
# #         except Exception:
# #             return None, None, None
        

# # class UpdateMemberInfoDB2(APIView):
# #     def post(self, request):
# #         name = request.data.get('name')
# #         relationship = request.data.get('relationship')
# #         ssn = request.data.get('ssn')
# #         member_id = request.data.get('member_id')
# #         dob_str = request.data.get('dob')  
# #         address = request.data.get('address')
# #         state = request.data.get('state')
# #         city = request.data.get('city')
# #         dep_ssn = request.data.get('dep_ssn',None)

# #         if not all([name, relationship, ssn, member_id, dob_str, address, state, city]):
# #             return Response({"error": "Missing required fields"}, status=400)

# #         year, month, day = self.parse_dob(dob_str)
# #         if not all([year, month, day]):
# #             return Response({"error": "Invalid date format. Expected format like 'mm-dd-yyyy'"}, status=400)

# #         try:
# #             connection = pyodbc.connect(connection_string
# #             )
# #             cursor = connection.cursor()

# #             if relationship.lower() == "member":
# #                 cursor.execute(f"SELECT * FROM {schema_name}.empyp WHERE EMSSN = ?", (ssn))
# #                 row = cursor.fetchone()
# #                 if not row:
# #                     return Response({"error": "Member record not found"}, status=404)

# #                 cursor.execute(f"""
# #                     UPDATE {schema_name}.empyp
# #                     SET EMNAME = ?, EMSSN = ?, EMMEM# = ?, EMDOBY = ?, EMDOBM = ?, EMDOBD = ?, EMADR1 = ?, EMST = ?, EMCITY = ?
# #                     WHERE EMSSN = ? WITH NC
# #                 """, (name, ssn, member_id, year, month, day, address, state, city, ssn))

# #             else:
# #                 cursor.execute(f"SELECT * FROM {schema_name}.depnp WHERE DPDSSN = ? " ,(dep_ssn,))
# #                 row = cursor.fetchone()
# #                 if not row:
# #                     return Response({"error": "Dependent record not found"}, status=404)

# #                 cursor.execute(f"""
# #                 UPDATE {schema_name}.depnp
# #                 SET "DPNAME" = ?, "DPDSSN" = ?, "DPDOBY" = ?, "DPDOBM" = ?, "DPDOBD" = ?
# #                 WHERE "DPDSSN" = ? WITH NC
# #             """, (name, dep_ssn, year, month, day, dep_ssn))

# #             connection.commit()
# #             cursor.close()
# #             connection.close()

# #             return Response({"message": "Record updated successfully"})

# #         except pyodbc.Error as e:
# #             return Response({"error": str(e)}, status=500)

# #     def parse_dob(self, dob_str):
# #         try:
# #             dt = datetime.strptime(dob_str, "%m-%d-%Y") 
# #             return dt.year, dt.month, dt.day
# #         except Exception:
# #             return None, None, None



# # class MostRecentDataView(APIView):
# #     def get(self, request):
# #         latest_entry = MyappRecentData.objects.order_by('-file_date').first()
# #         if latest_entry:
# #             latest_date = latest_entry.file_date
# #             latest_entries = MyappRecentData.objects.filter(file_date=latest_date)
# #             serializer = MyappRecentDataSerializer(latest_entries, many=True)
# #             return Response(serializer.data)
# #         return Response({"message": "No data found"}, status=404)
    

# # class AlternativeAddressTableCreate(APIView):
# #     def post(self, request, *args, **kwargs):
# #         serializer = AlternativeAddressTableSerializer(data=request.data)
        
# #         if serializer.is_valid():
# #             serializer.save()  
# #             return Response(serializer.data, status=status.HTTP_201_CREATED)
# #         return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

# # class AlternativeAddressTableUpdate(APIView):
# #     def post(self, request, *args, **kwargs):
# #         ssn = request.data.get('ssn', '')
# #         dep_ssn_list = request.data.get('dep_ssn', [])
# #         dep_name_list = request.data.get('dep_name', [])
# #         dep_relations_list = request.data.get('dep_relations', [])
# #         is_alternative = request.data.get('is_alternative', False)

# #         if is_alternative == "true":
# #             is_alternative = True

# #         if not ssn and not dep_ssn_list:
# #             return Response({"error": "SSN or DEPSSN is required to update the record"}, status=status.HTTP_400_BAD_REQUEST)

# #         if len(dep_ssn_list) != len(dep_name_list) or len(dep_ssn_list) != len(dep_relations_list):
# #             return Response({"error": "Mismatch in lengths of dep_ssn, dep_name, and dep_relations"}, status=status.HTTP_400_BAD_REQUEST)

# #         if is_alternative:
# #             instances = AlternativeAddressTable.objects.filter(ssn=ssn)
# #             if instances.count() > 1:
# #                 existing_dep_ssn_set = set(instances.values_list('dep_ssn', flat=True))
# #                 for dep_ssn, dep_name, dep_relation in zip(dep_ssn_list, dep_name_list, dep_relations_list):
# #                     record_data = request.data.copy()
# #                     record_data['dep_ssn'] = dep_ssn
# #                     record_data['last_first_name'] = dep_name
# #                     record_data['relationship'] = dep_relation

# #                     if str(dep_ssn) in existing_dep_ssn_set:
# #                         instance = AlternativeAddressTable.objects.get(ssn=ssn, dep_ssn=dep_ssn)
# #                         serializer = AlternativeAddressTableSerializer(instance, data=record_data, partial=False)
# #                         if serializer.is_valid():
# #                             serializer.save()
# #                     else:
# #                         serializer = AlternativeAddressTableSerializer(data=record_data)
# #                         if serializer.is_valid():
# #                             serializer.save()

# #             elif instances.count() == 1:
# #                 for dep_ssn, dep_name, dep_relation in zip(dep_ssn_list, dep_name_list, dep_relations_list):
# #                     new_record_data = request.data.copy()
# #                     new_record_data['dep_ssn'] = dep_ssn
# #                     new_record_data['last_first_name'] = dep_name
# #                     new_record_data['relationship'] = dep_relation
# #                     serializer = AlternativeAddressTableSerializer(data=new_record_data)
# #                     if serializer.is_valid():
# #                         serializer.save()

# #             elif not instances.exists():
# #                 print('great')
# #                 new_record = AlternativeAddressTable.objects.create(
# #                 last_first_name=request.data.get("last_first_name"),
# #                 pay_to_seq=request.data.get("pay_to_seq"),
# #                 address1=request.data.get("address1"),
# #                 address2=request.data.get("address2"),
# #                 address3=request.data.get("address3"),
# #                 city=request.data.get("city"),
# #                 state=request.data.get("state"),
# #                 zip=request.data.get("zip"),
# #                 relationship=request.data.get("relationship"),
# #                 last_activity_date=request.data.get("last_activity_date"),
# #                 employee_name=request.data.get("employee_name"),
# #                 ssn=ssn,
# #                 is_alternate_same=request.data.get("is_alternative"),
# #                 dep_ssn=" ")
            
# #                 for dep_ssn, dep_name, dep_relation in zip(dep_ssn_list, dep_name_list, dep_relations_list):
# #                     new_record_data = request.data.copy()
# #                     new_record_data['dep_ssn'] = dep_ssn
# #                     new_record_data['last_first_name'] = dep_name
# #                     new_record_data['relationship'] = dep_relation
# #                     serializer = AlternativeAddressTableSerializer(data=new_record_data)
# #                     if serializer.is_valid():
# #                         serializer.save()


# #             return Response({"message": "Records processed successfully"}, status=status.HTTP_200_OK)
# #         dep_flag = False
# #         if len(dep_ssn_list) == 1:
# #             dep_ssn = dep_ssn_list[0]
# #             instances = AlternativeAddressTable.objects.filter(dep_ssn=dep_ssn)
# #             dep_flag = True
# #         else:
# #             instances = AlternativeAddressTable.objects.filter(ssn=ssn, relationship="Member")

# #         if not instances.exists():
# #             return Response({"error": "No records found matching the criteria"}, status=status.HTTP_404_NOT_FOUND)

# #         updated_records = []
# #         print('hello',instances)
# #         for instance in instances:
# #             print("jake",instance)
# #             data = request.data.copy()
# #             print(data)
# #             if "dep_ssn" in data and data["dep_ssn"] is not None:
# #                 data["dep_ssn"] = str(data["dep_ssn"])

# #             if len(dep_ssn_list) == 0 and len(dep_name_list) == 0 and len(dep_relations_list) == 0:
# #                 serializer = AlternativeAddressTableSerializer(instance, data=data, partial=True)
# #                 if serializer.is_valid():
# #                     serializer.save()
# #                     updated_records.append(serializer.data)
# #                 else:
# #                     return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

# #             else:
# #                 for dep_ssn, dep_name, dep_relation in zip(dep_ssn_list, dep_name_list, dep_relations_list):
# #                     print('hell')
# #                     data["dep_ssn"] = dep_ssn
# #                     data["last_first_name"] = dep_name
# #                     data["relationship"] = dep_relation
# #                     print("help",instance)
# #                     serializer = AlternativeAddressTableSerializer(instance, data=data, partial=True)
# #                     if serializer.is_valid():
# #                         serializer.save()
# #                         updated_records.append(serializer.data)
# #                     else:
# #                         return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

# #         return Response({"message": "Records updated successfully", "updated_records": updated_records}, status=status.HTTP_200_OK)




# # class Get_AlternateAddressData(APIView):
# #     def get(self,request):
# #         dep_ssn = request.GET.get('dep_ssn','')
# #         ssn = request.GET.get('ssn','')
# #         total = request.GET.get('total',False)
# #         try:
# #             if total == 'True':
# #                 total = True
# #             if ssn and total:
# #                 data = list(AlternativeAddressTable.objects.filter(ssn=ssn).values())
# #                 return Response(data)
# #             elif ssn:
# #                 instance_data = AlternativeAddressTable.objects.filter(ssn=ssn, relationship="Member").first()
# #                 data = model_to_dict(instance_data) if instance_data else None
# #                 return Response(data)
# #             elif dep_ssn:
# #                 instance_data = AlternativeAddressTable.objects.get(dep_ssn=dep_ssn)
# #                 data = model_to_dict(instance_data) if instance_data else None
# #                 return Response(data)
# #             else:
# #                 return Response("SSN or DEP SSN is required")
# #         except:
# #             return Response("No record found for entered ssn")


# # class NotesEntryCreateView(APIView):
# #     def post(self, request, *args, **kwargs):
# #         serializer = NotesEntrySerializer(data=request.data)
# #         if serializer.is_valid():
# #             serializer.save()  
# #             return Response(serializer.data, status=status.HTTP_201_CREATED)
# #         return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    

# # class NotesEntryListView(APIView):
# #     def get(self, request, *args, **kwargs):
# #         notes_entries = NotesEntry.objects.all()
# #         serializer = NotesEntrySerializer(notes_entries, many=True)
# #         return Response(serializer.data, status=status.HTTP_200_OK)
    

# # class GetEligibilityData(APIView):
# #     def get(self, request):
# #         ssn = request.GET.get('ssn',None)
# #         dep_ssn = request.GET.get('dep_ssn',None)

# #         if not ssn or dep_ssn:
# #             return Response({"error": "SSN is required"}, status=status.HTTP_400_BAD_REQUEST)

# #         if ssn:
# #             try:
# #                 record = MyappEmpyp.objects.get(EMSSN=ssn)  # Fetch record based on SSN
# #             except MyappEmpyp.DoesNotExist:
# #                 return Response({"message": "No record found for the given SSN"}, status=status.HTTP_404_NOT_FOUND)

# #             # Define the specific fields you want to return
# #             response_data = {
# #                 "status date": record.EMEFFDATE,
# #                 "class code": record.EMCLAS,
# #                 "health status": record.EMFLAG
# #             }

# #             return Response(response_data, status=status.HTTP_200_OK)
        
# #         elif dep_ssn:
# #             try:
# #                 record = MyappDepnp.objects.get(DPDSSN=dep_ssn)  # Fetch record based on SSN
# #             except MyappDepnp.DoesNotExist:
# #                 return Response({"message": "No record found for the given DPDSSN"}, status=status.HTTP_404_NOT_FOUND)

# #             # Define the specific fields you want to return
# #             response_data = {
# #                 "status date": record.DPEFFDATE,
# #                 "class code": record.DPCLAS,
# #                 "health status": record.DPFLAG
# #             }

# #             return Response(response_data, status=status.HTTP_200_OK)


# # class GetEligibilityDataDB2(APIView):
# #     def get(self, request):
# #         ssn = request.GET.get('ssn', '')
# #         dep_ssn = request.GET.get('dep_ssn', '')
# #         mem_ssn = request.GET.get('mem_ssn','')
# #         dep_name = request.GET.get('name','')

# #         if not ssn and not dep_ssn:
# #             return Response({"error": "SSN or Dependent SSN is required"}, status=status.HTTP_400_BAD_REQUEST)

# #         # Database connection details
# #         host = '10.68.4.201'
# #         port = '23'
# #         database = 'S06e6f1r'
# #         user = 'ONEADMIN'
# #         password = 'ONEADMIN'

# #         connection_string = (
# #             f"DRIVER={{iSeries Access ODBC Driver}};"
# #             f"SYSTEM={host};"
# #             f"PORT={port};"
# #             f"DATABASE={database};"
# #             f"UID={user};"
# #             f"PWD={password};"
# #             f"PROTOCOL=TCPIP;"
# #         )

# #         try:
# #             connection = pyodbc.connect(connection_string)
# #             cursor = connection.cursor()

# #             response_data = {}

# #             if ssn:
# #                 query = f"""SELECT ELEPDY, ELEPDM, ELEPDD, ELPLAN, ELCLAS, ELUDTY, ELUDTM, ELUDTD, ELDSEQ, ELHSTA, ELUSER,ELWSTA
# #                         FROM {schema_name}.elghp 
# #                         WHERE ELSSN = ? AND ELDSEQ = ?"""
# #                 cursor.execute(query, (ssn, 0.0))
# #                 rows = cursor.fetchall()  # Fetch all matching records
# #                 print("lenght",len(rows))
# #                 if not rows:
# #                     return Response({"message": "No record found for the given SSN"}, status=status.HTTP_404_NOT_FOUND)

# #                 records = []
# #                 for row in rows:
# #                     elepdy, elepdm, elepdd, elplan, elclas, eludty, eludtm, eludtd,eldseq,elhsta,eluser,elwsta = row
# #                     emeffect_date = f"{str(elepdm).zfill(2)}/{str(elepdd).zfill(2)}/{str(elepdy).zfill(4)}"
# #                     last_updated_date = f"{str(eludtm).zfill(2)}/{str(eludtd).zfill(2)}/{str(eludty).zfill(4)}"

# #                     query = f"SELECT EMSTCD FROM {schema_name}.empyp WHERE EMSSN = ?"
# #                     cursor.execute(query, (ssn,))
# #                     health_status_row = cursor.fetchone()
# #                     health_status = health_status_row[0] if health_status_row else None

# #                     class_desc_list, current_class = get_class_name(str(elclas))
# #                     plan_desc = get_plan_name(str(elplan))
# #                     cob = check_COB(ssn)

# #                     records.append({
# #                         "status date": emeffect_date,  
# #                         "class code": elclas,  
# #                         "health status": health_status,  
# #                         "class desc list": class_desc_list,
# #                         "current class": current_class,
# #                         "plan desc": plan_desc,
# #                         "updated by": "user",
# #                         "last_updated_date": last_updated_date,
# #                         "eligibility_type": '',
# #                         "cob": cob,
# #                         "eldseq":eldseq,
# #                         "username":eluser,
# #                         "el_health_status":elhsta,
# #                         "weekly_status":elwsta
# #                     })

# #                 return Response({"records": records}, status=status.HTTP_200_OK)

# #             elif dep_ssn :
# #                 # seq_query = f"""SELECT DPSEQ FROM {schema_name}.depnp WHERE DPDSSN = ?"""
# #                 # cursor.execute(seq_query, (dep_ssn,))
# #                 # row = cursor.fetchone()
# #                 # dpseq = row[0]
# #                 # dpseq = float(dpseq)
# #                 # query = f"""SELECT ELEPDY, ELEPDM, ELEPDD, ELPLAN, ELCLAS, ELUDTY, ELUDTM, ELUDTD, ELDSEQ, ELHSTA, ELUSER,ELWSTA
# #                 #         FROM {schema_name}.elghp 
# #                 #         WHERE ELSSN = ? AND ELDSEQ = ?"""
# #                 # cursor.execute(query, (mem_ssn, dpseq))
# #                 # rows = cursor.fetchall()  

# #                 query = f"""
# #                     SELECT 
# #                         DPEFDY, DPEFDM, DPEFDD, DPPLAN, DPCLAS, DPUPYY, DPUPMM, DPUPDD,
# #                         DPSEQ, DPSTAT, DPUSER
# #                     FROM {schema_name}.depnp 
# #                     WHERE DPDSSN = ?
# #                 """
# #                 cursor.execute(query, (dep_ssn,))
# #                 rows = cursor.fetchall()

# #                 if not rows:
# #                     return Response({"message": "No record found for the given Dependent SSN"}, status=status.HTTP_404_NOT_FOUND)

# #                 records = []
# #                 for row in rows:
# #                     elepdy, elepdm, elepdd, elplan, elclas, eludty, eludtm, eludtd,eldseq,elhsta,eluser= row
# #                     emeffect_date = f"{str(elepdm).zfill(2)}/{str(elepdd).zfill(2)}/{str(elepdy).zfill(4)}"
# #                     last_updated_date = f"{str(eludtm).zfill(2)}/{str(eludtd).zfill(2)}/{str(eludty).zfill(4)}"

# #                     class_desc_list, current_class = get_class_name(str(elclas))
# #                     plan_desc = get_plan_name(str(elplan))
# #                     cob = check_COB(dep_ssn)

# #                     records.append({
# #                         "status date": emeffect_date,  
# #                         "class code": elclas,  
# #                         "health status": elhsta,  
# #                         "class desc list": class_desc_list,
# #                         "current class": current_class,
# #                         "plan desc": plan_desc,
# #                         "updated by": "user",
# #                         "last_updated_date": last_updated_date,
# #                         "eligibility_type": '',
# #                         "cob": cob,
# #                         "eldseq":eldseq,
# #                         "username":eluser,
# #                     })
# #                 return Response({"records": records}, status=status.HTTP_200_OK)


# #         except Exception as e:
# #             return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


# # class UpdateEligibilityDataDB2(APIView):
# #     def post(self, request):
# #         ssn = request.data.get("ssn", "")
# #         dep_ssn = request.data.get("dep_ssn", "")
# #         updated_data = request.data.get("updated_data", {})

# #         if not ssn and not dep_ssn:
# #             return Response({"error": "SSN or Dependent SSN is required"}, status=status.HTTP_400_BAD_REQUEST)

# #         try:
# #             connection = pyodbc.connect(connection_string)
# #             cursor = connection.cursor()

# #             # Extract date components from `effdate`
# #             effdate = updated_data.get("effdate", "")
# #             if effdate:
# #                 elepdm, elepdd, elepdy = effdate.split("/")  # Extract MM/DD/YYYY

# #             # Extract date components from `last_updated_date`
# #             last_updated_date = updated_data.get("last_updated_date", "")
# #             if last_updated_date:
# #                 eludtm, eludtd, eludty = last_updated_date.split("/")  # Extract MM/DD/YYYY

# #             if ssn:
# #                 insert_query = f"""
# #                     INSERT INTO {schema_name}.elghp (
# #                         ELSSN, ELEPDY, ELEPDM, ELEPDD, 
# #                         ELPLAN, ELCLAS, 
# #                         ELUDTY, ELUDTM, ELUDTD, ELHSTA
# #                     ) 
# #                     VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?) WITH NC
# #                 """
# #                 cursor.execute(insert_query, (
# #                     ssn, elepdy, elepdm, elepdd,
# #                     updated_data.get("elplan"),
# #                     updated_data.get("elclas"),
# #                     eludty, eludtm, eludtd,
# #                     updated_data.get("elhsta")
# #                 ))

# #                 update_stat_query = f"""
# #                     UPDATE {schema_name}.empyp
# #                     SET EMSTCD = ?, EMUPYY = ?,EMUPMM = ?,EMUPDD = ?
# #                     WHERE EMSSN = ? WITH NC
# #                 """
# #                 cursor.execute(update_stat_query,(updated_data.get('emstcd'),eludty,eludtm,eludtd,ssn))
            
# #             elif dep_ssn:
# #                 update_query = f"""
# #                     UPDATE {schema_name}.depnp 
# #                     SET DPEFDY = ?, DPEFDM = ?, DPEFDD = ?, 
# #                         DPCLAS = ?, DPPLAN = ?, DPSTAT = ?, 
# #                         DPUPYY = ?, DPUPMM = ?, DPUPDD = ? 
# #                     WHERE DPDSSN = ? WITH NC
# #                 """
# #                 cursor.execute(update_query, (
# #                     elepdy, elepdm, elepdd,
# #                     updated_data.get("dpclas"),
# #                     updated_data.get("dpplan"),
# #                     updated_data.get("dpstat"),
# #                     eludty, eludtm, eludtd,
# #                     dep_ssn
# #                 ))

# #             connection.commit()
# #             return Response({"message": "Record updated successfully"}, status=status.HTTP_200_OK)

# #         except Exception as e:
# #             return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

# #         finally:
# #             if cursor:
# #                 cursor.close()
# #             if connection:
# #                 connection.close()





# # import pyodbc
# # import os

# # port = '23'
# # host = '104.153.122.227'
# # database = 'S78F13CW'
# # user = 'ONENAGRJ'
# # password = 'Nagaraj8@'


# # def map_network_drive(drive_letter, network_path, username, password):
# #     try:
# #         # Command to map the network drive
# #         command = f'net use {drive_letter} {network_path} {password} /user:{username}'
# #         os.system(command)
# #         print(f"Network drive {drive_letter} mapped successfully.")
# #     except Exception as e:
# #         print(f"Error mapping network drive: {e}")


# # def call_stored_procedure_pdf(ssn, ssn_path, ssn_file):

   
# #     drive_letter = 'V:'
# #     network_path = r'\\104.153.122.227\HOME\DURGA'
# #     map_network_drive(drive_letter, network_path, user, password)


# #     connection_string = (
# #         f"DRIVER={{iSeries Access ODBC Driver}};" 
# #         f"SYSTEM={host};"
# #         f"PORT={port};"
# #         f"DATABASE={database};"
# #         f"UID={user};"
# #         f"PWD={password};"
# #         f"PROTOCOL=TCPIP;"
# #         f"CURRENTSCHEMA=QGPL;" 
# #     )

# #     param1 = ssn
# #     param2 = ssn_path
# #     param3 = ssn_file

# #     try:
# #         conn = pyodbc.connect(connection_string)
# #         cursor = conn.cursor()

# #         cursor.execute(
# #             "CALL QGPL.OOE_PROD_RUN_CL0164AS(?, ?, ?)", 
# #             (param1, param2, param3)
# #         )
# #         conn.commit()
# #         print("Stored procedure executed successfully.")
# #         return f"{drive_letter}\\{param3}"

# #     except pyodbc.Error as e:
# #         print(f"Error occurred: {e}")

# #     finally:
# #         if 'conn' in locals():
# #             cursor.close()
# #             conn.close()

# # @api_view(['GET'])
# # def generate_pdf(request):
# #     claim_no = request.GET.get('claim_no')
# #     ssn_path = "/HOME/DURGA"
# #     ssn_file = f"{claim_no}.pdf"
# #     print(claim_no)
# #     file_path = call_stored_procedure_pdf(claim_no, ssn_path, ssn_file)

# #     if file_path and os.path.exists(file_path):
# #         return FileResponse(open(file_path, "rb"), content_type="application/pdf")
    
# #     return JsonResponse({"error": "PDF file not found."}, status=404)



import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
import re

# Load the file
txt_file = r"C:\\Users\\krish\Downloads\\CA835_MW08905X_P_ERR_04030141.txt"
xlsx_file = "converted_report.xlsx"

# Read the lines
with open(txt_file, 'r') as file:
    lines = file.readlines()

# Find the index of header
header_index = None
for i, line in enumerate(lines):
    if "DCN" in line and "ITEM CODE" in line:
        header_index = i
        break

# Custom column headers
headers = [
    "DCN",
    "Item Code",
    "Total Charges",
    "Error Message",
    "Reject Code",
    "ID #",
    "Group ID",
    "Claim Submitter ID",
    "Member ID",
    "ITS-HME",
    "CLM-TYP",
    "COB",
    "Medicare"
]
from tabulate import tabulate
import os
# Extract only relevant data lines
data_lines = lines[header_index + 2:]
data_lines = [re.sub(r'\$(\d+),(\d+\.\d+)', r'$\1\2', line) for line in data_lines]
print(data_lines[0])
fields = re.split(r'(?<! ) {2}(?! )', data_lines[0].strip())
for dat in fields:
    print(dat)
rows = [re.split(r'\s{2,}', line) for line in data_lines]
df = pd.DataFrame(rows)
df.columns = headers
output_dir = 'output_files'  # You can change this to any desired directory
output_file = 'txt_to_excel.xlsx'

# Create the directory if it doesn't exist
os.makedirs(output_dir, exist_ok=True)

# Full path to the file
output_path = os.path.join(output_dir, output_file)

# Save the DataFrame to Excel
df.to_excel(output_path, index=False)

print(f"File saved successfully at: {output_path}")
# print(tabulate(rows, headers="firstrow", tablefmt="grid"))
# Split function using 2 or more spaces
# def split_line(line):
#     return re.split(r'\s{2,}', line.strip())

# # Parse the data safely
# data = []
# for line in data_lines:
#     if line.strip() == "":
#         continue  # Skip empty lines
#     row = split_line(line)
#     if len(row) == 16:
#         data.append(row)
#     else:
#        pass

# # Create DataFrame
# df = pd.DataFrame(data, columns=headers)

# # Save to Excel
# df.to_excel(xlsx_file, index=False)

# # Format Excel
# wb = load_workbook(xlsx_file)
# ws = wb.active

# # Bold headers & set alignment
# for cell in ws[1]:
#     cell.font = Font(bold=True)
#     cell.alignment = Alignment(horizontal="center")

# # Adjust column widths
# for col in ws.columns:
#     max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
#     col_letter = col[0].column_letter
#     ws.column_dimensions[col_letter].width = max_length + 2

# wb.save(xlsx_file)

print("✅ Conversion complete! Saved as:", xlsx_file)






